import io
import urllib.parse
import openpyxl
import asyncio
from pathlib import Path
from typing import Sequence, Callable
from datetime import date, datetime
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl.styles import Font, Alignment, Border, Side

from myapp.database.query_builders.query_case_filters import build_filtered_case_stmt
from myapp.models.repair_case_equipment import RepairCaseEquipment
from myapp.schemas import CaseFilterParams
from myapp.utils.export_utils import (
    format_serial_and_date,
    format_doc_number_and_date,
    format_section_mask,
)


class ExportService:

    TEMPLATE_PATH = Path("template_excel/template.xlsx")

    # Шрифт
    CELL_FONT = Font(name="Times New Roman", size=11, bold=True)

    # Выравнивание
    CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Границы
    CELL_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @staticmethod
    def _write_rows(
        ws, cases: Sequence[RepairCaseEquipment], start_row: int, row_builder: Callable
    ):
        """Метод для прохода по циклу и записи в таблицу"""

        for idx, case in enumerate(cases, start=start_row):
            row_values = row_builder(idx, case, start_row)

            for col_idx, value in enumerate(row_values, start=1):
                if isinstance(value, (date, datetime)):
                    value = value.strftime("%d.%m.%Y")

                cell = ws.cell(
                    row=idx, column=col_idx, value=value if value is not None else ""
                )

                cell.font = ExportService.CELL_FONT
                cell.alignment = ExportService.CELL_ALIGNMENT
                cell.border = ExportService.CELL_BORDER

            ws.row_dimensions[idx].height = None

    @staticmethod
    def _fill_main_sheet(ws, cases: Sequence[RepairCaseEquipment], start_row: int):
        def main_row_builder(idx, case, s_row):
            # Разматываем иерархию
            comp = case.component_equipment
            elem = case.element_equipment

            lvl2_name = comp.parent.equipment_name if (comp and comp.parent) else ""
            lvl3_name = comp.equipment_name if comp else ""
            lvl4_name = elem.equipment_name if elem else ""

            # Форматирование строк с серийниками (только если есть объект)
            old_comp_info = format_serial_and_date(
                case.component_quantity,
                case.component_serial_number_old,
                case.component_manufacture_date_old,
            )

            old_elem_info = format_serial_and_date(
                case.element_quantity,
                case.element_serial_number_old,
                case.element_manufacture_date_old,
            )

            # Новое оборудование (lvl3 или lvl4)
            new_eq_name = ""
            if case.new_element_equipment:
                new_eq_name = case.new_element_equipment.name
            elif case.new_component_equipment:
                new_eq_name = case.new_component_equipment.name

            new_eq_info = format_serial_and_date(
                case.new_component_quantity or case.new_element_quantity,
                case.component_serial_number_new or case.element_serial_number_new,
                case.component_manufacture_date_new
                or case.element_manufacture_date_new,
            )

            # Список значений согласно колонкам в таблице
            return [
                idx - s_row + 1,  # 1 №
                case.date_recorded,  # 2 Дата записи
                case.fault_date,  # 3 Дата неисправности
                case.locomotive_number,  # 4 Номер локомотива
                case.mileage,  # 5 Пробег
                (
                    case.locomotive_model.name if case.locomotive_model else ""
                ),  # 6 Модель локомотива
                format_section_mask(case.section_mask),  # 7 Секция
                (case.regional_center.name if case.regional_center else ""),  # 8 РЦ
                (
                    case.fault_discovered_at.name if case.fault_discovered_at else ""
                ),  # 9 Выявлено при
                lvl2_name,  # 10 Компонент (Lvl 2)
                lvl3_name,  # 11 Обозначение (Lvl 3)
                old_comp_info,  # 12 Зав. номер компонента
                lvl4_name,  # 13 Элемент (Lvl 4)
                old_elem_info,  # 14: Зав. номер элемента
                (case.malfunction.name if case.malfunction else ""),  # 15 Неисправность
                case.notes,  # 16 Примечание
                (case.repair_type.name if case.repair_type else ""),  # 17 Тип ремонта
                (
                    case.performed_by.name if case.performed_by else ""
                ),  # 18 Кем выполнен ремонт
                (
                    case.equipment_owner.name if case.equipment_owner else ""
                ),  # 19 Владелец нового оборудования
                (
                    case.destination.name if case.destination else ""
                ),  # 20 Куда отправили старое
                new_eq_name,  # 21 Вновь установленный
                new_eq_info,  # 22 Зав. номер нового
                (case.supplier.supplier_name if case.supplier else ""),  # 23 Поставщик
            ]

        ExportService._write_rows(ws, cases, start_row, main_row_builder)

    @staticmethod
    def _fill_warranty_sheet(ws, cases: Sequence[RepairCaseEquipment], start_row: int):
        def warranty_row_builder(idx, case, s_row):
            w = case.warranty_work

            return [
                idx - s_row + 1,  # 1 №
                w.notification_number if w else "",  # 2 № ув-ия
                w.notification_date if w else "",  # 3 Дата ув-ия
                (
                    w.notification_summary.name if w and w.notification_summary else ""
                ),  # 4 Содержание ув-ия
                (
                    format_doc_number_and_date(
                        w.re_notification_number, w.re_notification_date
                    )
                    if w
                    else ""
                ),  # 5 Повторное ув-ие (номер + дата)
                (
                    format_doc_number_and_date(
                        w.response_letter_number, w.response_letter_date
                    )
                    if w
                    else ""
                ),  # 6 Письмо-ответ (номер + дата)
                (
                    w.response_summary.name if w and w.response_summary else ""
                ),  # 7 Содержание ответа
                (
                    format_doc_number_and_date(w.claim_act_number, w.claim_act_date)
                    if w
                    else ""
                ),  # 8 РА (номер + дата)
                (
                    format_doc_number_and_date(
                        w.work_completion_act_number, w.work_completion_act_date
                    )
                    if w
                    else ""
                ),  # 9 АВР (номер + дата)
                (
                    w.decision_summary.name if w and w.decision_summary else ""
                ),  # 10 Решение
                (
                    w.research_status.name if w and w.research_status else ""
                ),  # 11 Статус исследования
                (
                    w.investigation_reason.name if w and w.investigation_reason else ""
                ),  # 12 Причина
                w.research_document if w else "",  # 13 Документ об исследовании
                (case.supplier.supplier_name if case.supplier else ""),  # 14 Поставщик
            ]

        ExportService._write_rows(ws, cases, start_row, warranty_row_builder)

    @staticmethod
    def _fill_waybill_sheet(ws, cases: Sequence[RepairCaseEquipment], start_row: int):
        def waybill_row_builder(idx, case, s_row):
            wb = case.waybill_doc

            return [
                idx - s_row + 1,  # 1 №
                (wb.ttn_replacement if wb else ""),  # 2 Документ об отпр. замены
                (wb.ttn_replacement_date if wb else ""),  # 3 Дата получения замены
                (wb.ttn_from_rc if wb else ""),  # 4 № ТТН из РЦ
                (wb.ttn_from_rc_date if wb else ""),  # 5 Дата поступления из РЦ
                (wb.ttn_to_supplier_date if wb else ""),  # 6 Дата отправки/списания
                (wb.ttn_to_supplier if wb else ""),  # 7 Документ отправки/списания
                (
                    wb.to_supplier_provider.name
                    if wb and wb.to_supplier_provider
                    else ""
                ),  # 8 Перевозчик
                (wb.ttn_from_supplier_date if wb else ""),  # 9 Дата возврата
                (wb.ttn_from_supplier if wb else ""),  # 10 ТТН возврата
                (
                    wb.from_supplier_provider.name
                    if wb and wb.from_supplier_provider
                    else ""
                ),  # 11 Перевозчик
                (case.supplier.supplier_name if case.supplier else ""),  # 12 Поставщик
            ]

        ExportService._write_rows(ws, cases, start_row, waybill_row_builder)

    @staticmethod
    def generate_cases_excel(cases: Sequence[RepairCaseEquipment]) -> io.BytesIO:
        """
        Генерирует Excel файл на основе шаблона, заполняя три разных листа.
        """
        if not ExportService.TEMPLATE_PATH.exists():
            raise FileNotFoundError(
                f"Шаблон не найден по пути: {ExportService.TEMPLATE_PATH}"
            )

        wb = openpyxl.load_workbook(ExportService.TEMPLATE_PATH)

        # Строка, с которой начинаются данные в твоем шаблоне (сразу после шапки)
        start_row = 2

        # Карта соответствия: Имя листа -> Функция для его заполнения
        sheets_to_fill = {
            "Рег. центр": ExportService._fill_main_sheet,
            "Рекл": ExportService._fill_warranty_sheet,
            "ТТН": ExportService._fill_waybill_sheet,
        }

        for sheet_name, fill_func in sheets_to_fill.items():
            if sheet_name in wb.sheetnames:
                fill_func(wb[sheet_name], cases, start_row)
            else:
                raise ValueError(f"Внимание: Лист '{sheet_name}' не найден в шаблоне")

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream

    @staticmethod
    async def get_cases_export_stream(
        session: AsyncSession, params: CaseFilterParams
    ) -> tuple[io.BytesIO, str]:
        """
        Достает данные из БД, генерирует Excel и возвращает кортеж:
        (поток с файлом, закодированное имя файла)
        """
        stmt = build_filtered_case_stmt(params, include_status=False)

        result = await session.execute(stmt)
        cases = result.unique().scalars().all()

        if not cases:
            raise ValueError("Нет данных для экспорта")

        file_stream = await asyncio.to_thread(ExportService.generate_cases_excel, cases)

        date_str = datetime.now().strftime("%d_%m_%Y")
        filename = f"Выгрузка_случаев_{date_str}.xlsx"
        encoded_filename = urllib.parse.quote(filename)

        return file_stream, encoded_filename
